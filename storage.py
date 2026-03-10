"""
storage.py — Layer di persistenza su file Excel per SigraFilm NOC.

Struttura cartella data/:
  data/utenti.xlsx        — utenti
  data/cinema.xlsx        — cinema
  data/cinema_eliminati.xlsx — cinema eliminati (tombstone)
  data/tickets.xlsx       — ticket (aperti + chiusi)
  data/commenti.xlsx      — commenti ai ticket
  data/letture.xlsx       — tracciamento lettura ticket per utente
  data/assegnazioni.xlsx  — assegnazioni utente→cinema
"""

import os
import threading
from datetime import datetime
from dataclasses import dataclass, field
from typing import List, Optional, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash

# ─── Percorso cartella dati ──────────────────────────
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

# Lock globale — protegge da scritture concorrenti
_lock = threading.Lock()


# ══════════════════════════════════════════════════════
# DATACLASSES
# ══════════════════════════════════════════════════════

@dataclass
class User:
    id: int
    username: str
    password_hash: str
    password_plain: str = ""
    role: str = "user"
    telefono: str = ""
    email: str = ""


@dataclass
class Problem:
    id: int
    cinema: str
    città: str
    sala: str
    tipo: str
    urgenza: str
    stato: str = "Aperto"
    chiuso_da: Optional[str] = None
    chiuso_il: Optional[datetime] = None
    autore: str = ""
    data_ora: Optional[datetime] = None

    def __post_init__(self):
        if self.data_ora is None:
            self.data_ora = datetime.utcnow()

    @property
    def comments(self):
        return store.get_comments(self.id)


@dataclass
class Comment:
    id: int
    problem_id: int
    autore: str
    role: str
    testo: str
    data_ora: Optional[datetime] = None

    def __post_init__(self):
        if self.data_ora is None:
            self.data_ora = datetime.utcnow()


@dataclass
class Cinema:
    id: int
    nome: str
    città: str = ""
    num_sale: int = 1
    telefono: str = ""
    indirizzo: str = ""
    lat: Optional[float] = None
    lng: Optional[float] = None


@dataclass
class TicketRead:
    id: int
    user_id: int
    problem_id: int
    last_read_at: Optional[datetime] = None

    def __post_init__(self):
        if self.last_read_at is None:
            self.last_read_at = datetime.utcnow()


@dataclass
class UserCinema:
    id: int
    user_id: int
    cinema_id: int


# ══════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════

def _ensure_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


def _path(name: str) -> str:
    return os.path.join(DATA_DIR, name)


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def _i(v, default=0) -> int:
    try:
        return int(v) if v is not None else default
    except (ValueError, TypeError):
        return default


def _f(v) -> Optional[float]:
    try:
        return float(v) if v is not None and str(v).strip() != "" else None
    except (ValueError, TypeError):
        return None


def _dt(v) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _fmt_dt(dt: Optional[datetime]) -> str:
    return dt.strftime("%d/%m/%Y %H:%M") if dt else ""


def _load_wb(filename: str, headers: list) -> openpyxl.Workbook:
    """Carica o crea un workbook con la riga header se non esiste."""
    path = _path(filename)
    if os.path.exists(path):
        try:
            return openpyxl.load_workbook(path)
        except Exception:
            pass
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_header(ws, headers)
    wb.save(path)
    return wb


def _write_header(ws, headers: list):
    ws.append(headers)
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")


def _save_wb(wb: openpyxl.Workbook, filename: str):
    wb.save(_path(filename))


def _next_id(rows: list) -> int:
    if not rows:
        return 1
    return max((_i(r[0], 0) for r in rows), default=0) + 1


# ══════════════════════════════════════════════════════
# STORE — classe principale
# ══════════════════════════════════════════════════════

class ExcelStore:
    """
    Accesso completo ai dati tramite file Excel.
    Ogni metodo acquisisce il lock prima di leggere/scrivere.
    """

    HEADERS = {
        "utenti.xlsx":           ["id", "username", "password_hash", "password_plain", "role", "telefono", "email"],
        "cinema.xlsx":           ["id", "nome", "città", "num_sale", "telefono", "indirizzo", "lat", "lng"],
        "cinema_eliminati.xlsx": ["id", "nome"],
        "tickets.xlsx":          ["id", "cinema", "città", "sala", "tipo", "urgenza", "stato",
                                  "chiuso_da", "chiuso_il", "autore", "data_ora"],
        "commenti.xlsx":         ["id", "problem_id", "autore", "role", "testo", "data_ora"],
        "letture.xlsx":          ["id", "user_id", "problem_id", "last_read_at"],
        "assegnazioni.xlsx":     ["id", "user_id", "cinema_id"],
    }

    def _rows(self, filename: str) -> list:
        """Restituisce le righe (senza header) come lista di tuple."""
        wb = _load_wb(filename, self.HEADERS[filename])
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        return rows[1:] if len(rows) > 1 else []

    def _overwrite(self, filename: str, data_rows: list):
        """Riscrive il file Excel da zero con header + righe dati."""
        _ensure_dir()
        wb = openpyxl.Workbook()
        ws = wb.active
        _write_header(ws, self.HEADERS[filename])
        for row in data_rows:
            ws.append(row)
        _save_wb(wb, filename)

    # ── USERS ─────────────────────────────────────────

    def _row_to_user(self, r) -> User:
        return User(
            id=_i(r[0]), username=_s(r[1]), password_hash=_s(r[2]),
            password_plain=_s(r[3]), role=_s(r[4]) or "user",
            telefono=_s(r[5]), email=_s(r[6]),
        )

    def _user_to_row(self, u: User) -> tuple:
        return (u.id, u.username, u.password_hash, u.password_plain,
                u.role, u.telefono, u.email)

    def get_all_users(self) -> List[User]:
        with _lock:
            return [self._row_to_user(r) for r in self._rows("utenti.xlsx")]

    def get_user_by_id(self, uid: int) -> Optional[User]:
        with _lock:
            for r in self._rows("utenti.xlsx"):
                if _i(r[0]) == uid:
                    return self._row_to_user(r)
        return None

    def get_user_by_username(self, username: str) -> Optional[User]:
        with _lock:
            for r in self._rows("utenti.xlsx"):
                if _s(r[1]).lower() == username.lower():
                    return self._row_to_user(r)
        return None

    def create_user(self, username: str, password_hash: str, password_plain: str = "",
                    role: str = "user", telefono: str = "", email: str = "") -> User:
        with _lock:
            rows = self._rows("utenti.xlsx")
            new_id = _next_id(rows)
            u = User(id=new_id, username=username, password_hash=password_hash,
                     password_plain=password_plain, role=role, telefono=telefono, email=email)
            rows.append(self._user_to_row(u))
            self._overwrite("utenti.xlsx", rows)
            return u

    def update_user(self, u: User):
        with _lock:
            rows = self._rows("utenti.xlsx")
            rows = [self._user_to_row(u) if _i(r[0]) == u.id else r for r in rows]
            self._overwrite("utenti.xlsx", rows)

    def delete_user(self, uid: int):
        with _lock:
            rows = [r for r in self._rows("utenti.xlsx") if _i(r[0]) != uid]
            self._overwrite("utenti.xlsx", rows)
            # rimuovi assegnazioni
            arows = [r for r in self._rows("assegnazioni.xlsx") if _i(r[1]) != uid]
            self._overwrite("assegnazioni.xlsx", arows)

    def count_admins(self) -> int:
        return sum(1 for u in self.get_all_users() if u.role == "admin")

    # ── PROBLEMS ──────────────────────────────────────

    def _row_to_problem(self, r) -> Problem:
        return Problem(
            id=_i(r[0]), cinema=_s(r[1]), città=_s(r[2]), sala=_s(r[3]) or "1",
            tipo=_s(r[4]), urgenza=_s(r[5]), stato=_s(r[6]) or "Aperto",
            chiuso_da=_s(r[7]) or None,
            chiuso_il=_dt(r[8]),
            autore=_s(r[9]),
            data_ora=_dt(r[10]) or datetime.utcnow(),
        )

    def _problem_to_row(self, p: Problem) -> tuple:
        return (p.id, p.cinema, p.città, p.sala, p.tipo, p.urgenza, p.stato,
                p.chiuso_da or "", _fmt_dt(p.chiuso_il), p.autore, _fmt_dt(p.data_ora))

    def get_all_problems(self) -> List[Problem]:
        with _lock:
            return [self._row_to_problem(r) for r in self._rows("tickets.xlsx")]

    def get_problem_by_id(self, pid: int) -> Optional[Problem]:
        with _lock:
            for r in self._rows("tickets.xlsx"):
                if _i(r[0]) == pid:
                    return self._row_to_problem(r)
        return None

    def get_problems_filtered(self, stato_ne: str = None, stato_eq: str = None,
                               autore: str = None, urgenza: str = None) -> List[Problem]:
        problems = self.get_all_problems()
        if stato_ne:
            problems = [p for p in problems if p.stato != stato_ne]
        if stato_eq:
            problems = [p for p in problems if p.stato == stato_eq]
        if autore:
            problems = [p for p in problems if p.autore == autore]
        if urgenza:
            problems = [p for p in problems if p.urgenza == urgenza]
        return sorted(problems, key=lambda p: p.data_ora or datetime.min, reverse=True)

    def create_problem(self, cinema: str, città: str, sala: str, tipo: str,
                       urgenza: str, stato: str, autore: str) -> Problem:
        with _lock:
            rows = self._rows("tickets.xlsx")
            new_id = _next_id(rows)
            p = Problem(id=new_id, cinema=cinema, città=città, sala=sala,
                        tipo=tipo, urgenza=urgenza, stato=stato, autore=autore,
                        data_ora=datetime.utcnow())
            rows.append(self._problem_to_row(p))
            self._overwrite("tickets.xlsx", rows)
            return p

    def update_problem(self, p: Problem):
        with _lock:
            rows = self._rows("tickets.xlsx")
            rows = [self._problem_to_row(p) if _i(r[0]) == p.id else r for r in rows]
            self._overwrite("tickets.xlsx", rows)

    def delete_problem(self, pid: int):
        with _lock:
            rows = [r for r in self._rows("tickets.xlsx") if _i(r[0]) != pid]
            self._overwrite("tickets.xlsx", rows)
            # rimuovi commenti e letture associati
            crows = [r for r in self._rows("commenti.xlsx") if _i(r[1]) != pid]
            self._overwrite("commenti.xlsx", crows)
            lrows = [r for r in self._rows("letture.xlsx") if _i(r[2]) != pid]
            self._overwrite("letture.xlsx", lrows)

    # ── COMMENTS ──────────────────────────────────────

    def _row_to_comment(self, r) -> Comment:
        return Comment(
            id=_i(r[0]), problem_id=_i(r[1]), autore=_s(r[2]),
            role=_s(r[3]) or "user", testo=_s(r[4]),
            data_ora=_dt(r[5]) or datetime.utcnow(),
        )

    def _comment_to_row(self, c: Comment) -> tuple:
        return (c.id, c.problem_id, c.autore, c.role, c.testo, _fmt_dt(c.data_ora))

    def get_comments(self, problem_id: int) -> List[Comment]:
        with _lock:
            rows = [r for r in self._rows("commenti.xlsx") if _i(r[1]) == problem_id]
            comments = [self._row_to_comment(r) for r in rows]
            return sorted(comments, key=lambda c: c.data_ora or datetime.min)

    def add_comment(self, problem_id: int, autore: str, role: str, testo: str) -> Comment:
        with _lock:
            rows = self._rows("commenti.xlsx")
            new_id = _next_id(rows)
            c = Comment(id=new_id, problem_id=problem_id, autore=autore,
                        role=role, testo=testo, data_ora=datetime.utcnow())
            rows.append(self._comment_to_row(c))
            self._overwrite("commenti.xlsx", rows)
            return c

    # ── CINEMAS ───────────────────────────────────────

    def _row_to_cinema(self, r) -> Cinema:
        return Cinema(
            id=_i(r[0]), nome=_s(r[1]), città=_s(r[2]),
            num_sale=_i(r[3], 1), telefono=_s(r[4]),
            indirizzo=_s(r[5]), lat=_f(r[6]), lng=_f(r[7]),
        )

    def _cinema_to_row(self, c: Cinema) -> tuple:
        return (c.id, c.nome, c.città, c.num_sale, c.telefono,
                c.indirizzo, c.lat or "", c.lng or "")

    def get_all_cinemas(self, order_by: str = "nome") -> List[Cinema]:
        with _lock:
            cinemas = [self._row_to_cinema(r) for r in self._rows("cinema.xlsx")]
        if order_by == "città_nome":
            cinemas.sort(key=lambda c: (c.città, c.nome))
        else:
            cinemas.sort(key=lambda c: c.nome)
        return cinemas

    def get_cinema_by_id(self, cid: int) -> Optional[Cinema]:
        with _lock:
            for r in self._rows("cinema.xlsx"):
                if _i(r[0]) == cid:
                    return self._row_to_cinema(r)
        return None

    def get_cinemas_by_ids(self, ids: list) -> List[Cinema]:
        with _lock:
            return [self._row_to_cinema(r)
                    for r in self._rows("cinema.xlsx") if _i(r[0]) in ids]

    def get_cinema_by_nome(self, nome: str) -> Optional[Cinema]:
        with _lock:
            for r in self._rows("cinema.xlsx"):
                if _s(r[1]) == nome:
                    return self._row_to_cinema(r)
        return None

    def create_cinema(self, nome: str, città: str, num_sale: int = 1,
                      telefono: str = "", indirizzo: str = "",
                      lat=None, lng=None) -> Cinema:
        with _lock:
            rows = self._rows("cinema.xlsx")
            new_id = _next_id(rows)
            c = Cinema(id=new_id, nome=nome, città=città, num_sale=num_sale,
                       telefono=telefono, indirizzo=indirizzo, lat=lat, lng=lng)
            rows.append(self._cinema_to_row(c))
            self._overwrite("cinema.xlsx", rows)
            return c

    def update_cinema(self, c: Cinema):
        with _lock:
            rows = self._rows("cinema.xlsx")
            rows = [self._cinema_to_row(c) if _i(r[0]) == c.id else r for r in rows]
            self._overwrite("cinema.xlsx", rows)

    def delete_cinema(self, cid: int):
        with _lock:
            cinema = None
            for r in self._rows("cinema.xlsx"):
                if _i(r[0]) == cid:
                    cinema = self._row_to_cinema(r)
                    break
            rows = [r for r in self._rows("cinema.xlsx") if _i(r[0]) != cid]
            self._overwrite("cinema.xlsx", rows)
            if cinema:
                self._add_deleted_cinema(cinema.nome)
            # rimuovi assegnazioni
            arows = [r for r in self._rows("assegnazioni.xlsx") if _i(r[2]) != cid]
            self._overwrite("assegnazioni.xlsx", arows)

    def get_all_cinema_nomi(self) -> set:
        with _lock:
            return {_s(r[1]) for r in self._rows("cinema.xlsx")}

    # ── CINEMA ELIMINATI ──────────────────────────────

    def _add_deleted_cinema(self, nome: str):
        rows = self._rows("cinema_eliminati.xlsx")
        if not any(_s(r[1]) == nome for r in rows):
            new_id = _next_id(rows)
            rows.append((new_id, nome))
            self._overwrite("cinema_eliminati.xlsx", rows)

    def get_deleted_cinema_nomi(self) -> set:
        with _lock:
            return {_s(r[1]) for r in self._rows("cinema_eliminati.xlsx")}

    # ── TICKET READS ──────────────────────────────────

    def _row_to_ticketread(self, r) -> TicketRead:
        return TicketRead(
            id=_i(r[0]), user_id=_i(r[1]), problem_id=_i(r[2]),
            last_read_at=_dt(r[3]),
        )

    def _tr_to_row(self, tr: TicketRead) -> tuple:
        return (tr.id, tr.user_id, tr.problem_id, _fmt_dt(tr.last_read_at))

    def get_reads_by_user(self, user_id: int) -> Dict[int, datetime]:
        """Restituisce {problem_id: last_read_at} per l'utente."""
        with _lock:
            result = {}
            for r in self._rows("letture.xlsx"):
                if _i(r[1]) == user_id:
                    result[_i(r[2])] = _dt(r[3])
            return result

    def upsert_ticket_read(self, user_id: int, problem_id: int):
        with _lock:
            rows = self._rows("letture.xlsx")
            now_str = _fmt_dt(datetime.utcnow())
            found = False
            new_rows = []
            for r in rows:
                if _i(r[1]) == user_id and _i(r[2]) == problem_id:
                    new_rows.append((r[0], user_id, problem_id, now_str))
                    found = True
                else:
                    new_rows.append(r)
            if not found:
                new_id = _next_id(rows)
                new_rows.append((new_id, user_id, problem_id, now_str))
            self._overwrite("letture.xlsx", new_rows)

    # ── USER-CINEMA ASSIGNMENTS ────────────────────────

    def get_cinema_ids_for_user(self, user_id: int) -> List[int]:
        with _lock:
            return [_i(r[2]) for r in self._rows("assegnazioni.xlsx")
                    if _i(r[1]) == user_id]

    def set_user_cinemas(self, user_id: int, cinema_ids: List[int]):
        with _lock:
            rows = [r for r in self._rows("assegnazioni.xlsx") if _i(r[1]) != user_id]
            base_id = _next_id(rows) if rows else 1
            for i, cid in enumerate(cinema_ids):
                rows.append((base_id + i, user_id, cid))
            self._overwrite("assegnazioni.xlsx", rows)

    # ── SEED ──────────────────────────────────────────

    def seed(self):
        """Inizializza i file e inserisce dati di default se mancanti."""
        _ensure_dir()
        # Assicura che tutti i file esistano
        for fname, headers in self.HEADERS.items():
            if not os.path.exists(_path(fname)):
                _load_wb(fname, headers)

        # Admin di default
        if not self.get_user_by_username("admin"):
            self.create_user(
                username="admin",
                password_hash=generate_password_hash("admin1234"),
                password_plain="admin1234",
                role="admin",
            )
            print("✅ Utente admin creato (admin / admin1234)")

        # Seed cinema
        existing_nomi = self.get_all_cinema_nomi()
        deleted_nomi = self.get_deleted_cinema_nomi()
        _cinemas_seed = [
            {"nome": "Cinema Chiusi",                        "città": "Chiusi",                   "num_sale": 6, "telefono": "0578 275077", "indirizzo": "Loc. Querce al Pino, SP 146, 53043 Chiusi SI",         "lat": 43.0025, "lng": 11.9481},
            {"nome": "Cinema Empoli",                        "città": "Empoli",                   "num_sale": 3, "telefono": "0571 72023",  "indirizzo": "Via Cosimo Ridolfi 75, 50053 Empoli FI",              "lat": 43.7208, "lng": 10.9478},
            {"nome": "Cinema Firenze",                       "città": "Firenze",                  "num_sale": 1, "telefono": "055 483607",  "indirizzo": "Via G. Romagnosi 46, 50134 Firenze FI",               "lat": 43.7835, "lng": 11.2427},
            {"nome": "Cinema Odeon",                         "città": "Firenze",                  "num_sale": 1, "telefono": "055 214068",  "indirizzo": "Piazza degli Strozzi 2, 50123 Firenze FI",            "lat": 43.7711, "lng": 11.2519},
            {"nome": "Cinema Grosseto",                      "città": "Grosseto",                 "num_sale": 4, "telefono": "0564 27069",  "indirizzo": "Via Goffredo Mameli 24, 58100 Grosseto GR",           "lat": 42.7641, "lng": 11.1086},
            {"nome": "Cinema Massa",                         "città": "Massa",                    "num_sale": 7, "telefono": "0585 791105", "indirizzo": "Via Dorsale 11, 54100 Massa MS",                      "lat": 44.0181, "lng": 10.1327},
            {"nome": "Cinema Montecatini",                   "città": "Montecatini Terme",        "num_sale": 4, "telefono": "0572 78510",  "indirizzo": "Piazza Massimo D'Azeglio 5, 51016 Montecatini Terme PT", "lat": 43.8849, "lng": 10.7722},
            {"nome": "Cinema Pisa",                          "città": "Pisa",                     "num_sale": 3, "telefono": "050 5552261", "indirizzo": "Via Piave 47, 56123 Pisa PI",                         "lat": 43.7155, "lng": 10.3986},
            {"nome": "Cinecity Pisa",                        "città": "Pisa",                     "num_sale": 5, "telefono": "392 323 3535","indirizzo": "Piazza della Stazione 16, 56125 Pisa PI",             "lat": 43.7090, "lng": 10.3972},
            {"nome": "Cinema Sansepolcro",                   "città": "Sansepolcro",              "num_sale": 1, "telefono": "0575 733433", "indirizzo": "Via XX Settembre 156, 52037 Sansepolcro AR",           "lat": 43.5695, "lng": 12.1406},
            {"nome": "ELIA ANTICA MULTISALA",                "città": "Grosseto",                 "num_sale": 4, "telefono": "0564 644987", "indirizzo": "Via Aurelia Antica 46, 58100 Grosseto GR",            "lat": 42.7548, "lng": 11.0931},
            {"nome": "Cinema Scuderie Granducali Seravezza", "città": "Seravezza",                "num_sale": 1, "telefono": "0584 840409", "indirizzo": "Viale Leonetto Amedei 124, 55047 Seravezza LU",       "lat": 43.9962, "lng": 10.2321},
            {"nome": "Teatro Cinema Giotto",                 "città": "Borgo San Lorenzo",        "num_sale": 1, "telefono": "055 845 9658","indirizzo": "Corso Giacomo Matteotti 151, 50032 Borgo San Lorenzo FI", "lat": 43.9548, "lng": 11.3855},
            {"nome": "Cinema Metropolitan",                  "città": "Piombino",                 "num_sale": 1, "telefono": "0565 30385",  "indirizzo": "Piazza Cappelletti 2, 57025 Piombino LI",             "lat": 42.9225, "lng": 10.5320},
            {"nome": "Cinema Multisala Excelsior",           "città": "Montecatini Terme",        "num_sale": 2, "telefono": "0572 904289", "indirizzo": "Viale Giuseppe Verdi 66, 51016 Montecatini Terme PT", "lat": 43.8825, "lng": 10.7740},
            {"nome": "Cinema Teatro Scipione Ammirato",      "città": "Montaione",                "num_sale": 1, "telefono": "0571 61517",  "indirizzo": "Piazza Gramsci 2, 50050 Montaione FI",                "lat": 43.5595, "lng": 10.9126},
            {"nome": "Multisala Isola Verde",                "città": "Pisa",                     "num_sale": 3, "telefono": "050 973676",  "indirizzo": "Via Vittorio Frascani, 56124 Pisa PI",                "lat": 43.7024, "lng": 10.3912},
            {"nome": "Cinema Sala Esse",                     "città": "Firenze",                  "num_sale": 1, "telefono": "055 666643",  "indirizzo": "Via del Ghirlandaio 38, 50121 Firenze FI",            "lat": 43.7697, "lng": 11.2763},
            {"nome": "Multisala Goldoni",                    "città": "Viareggio",                "num_sale": 2, "telefono": "0584 49832",  "indirizzo": "Via San Francesco 124, 55049 Viareggio LU",           "lat": 43.8682, "lng": 10.2547},
            {"nome": "Cinema Multisala Il Portico",          "città": "Firenze",                  "num_sale": 2, "telefono": "055 669930",  "indirizzo": "Via Capo di Mondo 66, 50136 Firenze FI",              "lat": 43.7698, "lng": 11.2919},
            {"nome": "Cinema Teatro Everest Galluzzo",       "città": "Firenze",                  "num_sale": 1, "telefono": "055 232 1754","indirizzo": "Via Volterrana 4, 50124 Firenze FI",                  "lat": 43.7388, "lng": 11.2413},
            {"nome": "Spazio Alfieri Cinema Teatro Bistrò",  "città": "Firenze",                  "num_sale": 1, "telefono": "055 5320840", "indirizzo": "Via dell'Ulivo 8, 50122 Firenze FI",                  "lat": 43.7703, "lng": 11.2639},
            {"nome": "Cinema Teatro Multisala Imperiale",    "città": "Montecatini Terme",        "num_sale": 4, "telefono": "0572 508601", "indirizzo": "Piazza Massimo D'Azeglio 5, 51016 Montecatini Terme PT", "lat": 43.8849, "lng": 10.7722},
            {"nome": "Cinema Centrale",                      "città": "Viareggio",                "num_sale": 1, "telefono": "0584 581226", "indirizzo": "Via Cesare Battisti 67, 55049 Viareggio LU",          "lat": 43.8707, "lng": 10.2534},
            {"nome": "Cinema Nuova Aurora",                  "città": "Sansepolcro",              "num_sale": 1, "telefono": "0575 1480629","indirizzo": "Via Piero della Francesca 47, 52037 Sansepolcro AR",  "lat": 43.5696, "lng": 12.1393},
            {"nome": "Cinema Marconi",                       "città": "Firenze",                  "num_sale": 3, "telefono": "055 680554",  "indirizzo": "Viale Giannotti 45r, 50126 Firenze FI",               "lat": 43.7526, "lng": 11.2694},
            {"nome": "Multisala Splendor",                   "città": "Massa",                    "num_sale": 7, "telefono": "0585 791105", "indirizzo": "Via Dorsale 11, 54100 Massa MS",                      "lat": 44.0181, "lng": 10.1327},
            {"nome": "Teatro dei Servi",                     "città": "Massa",                    "num_sale": 1, "telefono": "0585 811973", "indirizzo": "Via Palestro 37, 54100 Massa MS",                     "lat": 44.0300, "lng": 10.1406},
            {"nome": "Multisala Odeon",                      "città": "Pisa",                     "num_sale": 4, "telefono": "050 540168",  "indirizzo": "Piazza S. Paolo all'Orto 18, 56127 Pisa PI",          "lat": 43.7188, "lng": 10.4040},
            {"nome": "Cinema Caffè Lanteri",                 "città": "Pisa",                     "num_sale": 1, "telefono": "050 577100",  "indirizzo": "Via San Michele degli Scalzi 46, 56124 Pisa PI",      "lat": 43.7188, "lng": 10.4180},
            {"nome": "Cinema Teatro 4 Mori",                 "città": "Livorno",                  "num_sale": 1, "telefono": "342 543 1247","indirizzo": "Via Pietro Tacca 16, 57123 Livorno LI",               "lat": 43.5498, "lng": 10.3122},
            {"nome": "Multisala Eden",                       "città": "Arezzo",                   "num_sale": 2, "telefono": "0575 353364", "indirizzo": "Via Antonio Guadagnoli 2, 52100 Arezzo AR",            "lat": 43.4632, "lng": 11.8792},
            {"nome": "Nuovo Cinema Caporali",                "città": "Castiglione del Lago",     "num_sale": 3, "telefono": "075 965 3152","indirizzo": "Piazzetta San Domenico 1, 06061 Castiglione del Lago PG", "lat": 43.1200, "lng": 12.0557},
            {"nome": "Cinema Teatro Verdi",                  "città": "San Vincenzo",             "num_sale": 1, "telefono": "0565 701918", "indirizzo": "Via Vittorio Emanuele II 121, 57027 San Vincenzo LI",  "lat": 43.0990, "lng": 10.5398},
            {"nome": "Teatro Signorelli",                    "città": "Cortona",                  "num_sale": 1, "telefono": "0575 601882", "indirizzo": "Piazza Signorelli 13, 52044 Cortona AR",               "lat": 43.2763, "lng": 11.9876},
            {"nome": "Cinema Città di Villafranca",          "città": "Villafranca in Lunigiana", "num_sale": 1, "telefono": "0187 498011", "indirizzo": "Via Roma 2, 54028 Villafranca in Lunigiana MS",        "lat": 44.3035, "lng":  9.9536},
            {"nome": "Cinema Teatro Excelsior",              "città": "Reggello",                 "num_sale": 1, "telefono": "055 869190",  "indirizzo": "Via Dante Alighieri 7, 50066 Reggello FI",            "lat": 43.6845, "lng": 11.5340},
            {"nome": "Cinema Arena Ardenza",                 "città": "Livorno",                  "num_sale": 1, "telefono": "0586 501403", "indirizzo": "Piazza Sforzini 17, 57128 Livorno LI",                "lat": 43.4980, "lng": 10.3350},
            {"nome": "Arena Dentro Le Mura",                 "città": "San Casciano Val di Pesa", "num_sale": 1, "telefono": "",            "indirizzo": "Via Lucardesi 10, 50026 San Casciano Val di Pesa FI", "lat": 43.6563, "lng": 11.1832},
        ]
        added = 0
        for c in _cinemas_seed:
            if c["nome"] not in existing_nomi and c["nome"] not in deleted_nomi:
                self.create_cinema(**c)
                existing_nomi.add(c["nome"])
                added += 1
        if added:
            print(f"✅ {added} cinema aggiunti al catalogo")

    # ── IMPORT da Excel esterno ────────────────────────

    def import_from_workbook(self, wb) -> dict:
        """
        Importa ticket e cinema da un workbook openpyxl caricato dall'utente.
        Restituisce contatori: added_problems, skipped_problems, added_cinemas, skipped_cinemas.
        """
        added_problems = 0
        skipped_problems = 0
        added_cinemas = 0
        skipped_cinemas = 0

        existing_ids = {p.id for p in self.get_all_problems()}
        existing_nomi = self.get_all_cinema_nomi()

        def _val(v):
            return str(v).strip() if v is not None else ""

        def _parse_dt_imp(val):
            if not val:
                return None
            if isinstance(val, datetime):
                return val
            return _dt(val)

        for sheet_name in ["Ticket Aperti", "Archivio Chiusi"]:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if not ws:
                continue
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            for row in rows[1:]:
                if not any(row):
                    continue
                try:
                    row_id  = int(row[0]) if row[0] else None
                    cinema  = _val(row[1])
                    città   = _val(row[2])
                    sala    = _val(row[3]) or "1"
                    tipo    = _val(row[4])
                    urgenza = _val(row[5]) or "Non urgente"
                    autore  = _val(row[7]) if len(row) > 7 else "import"
                    data_ora = _parse_dt_imp(row[8]) if len(row) > 8 else None
                    stato   = _val(row[6]) if sheet_name == "Ticket Aperti" else "Chiuso"
                    chiuso_da = _val(row[9]) if len(row) > 9 else None
                    chiuso_il = _parse_dt_imp(row[10]) if len(row) > 10 else None
                except Exception:
                    continue
                if not cinema or not tipo:
                    continue
                if row_id and row_id in existing_ids:
                    skipped_problems += 1
                    continue
                p = self.create_problem(cinema=cinema, città=città, sala=sala,
                                        tipo=tipo, urgenza=urgenza, stato=stato, autore=autore)
                if data_ora:
                    p.data_ora = data_ora
                p.chiuso_da = chiuso_da or None
                p.chiuso_il = chiuso_il
                self.update_problem(p)
                if row_id:
                    existing_ids.add(row_id)
                added_problems += 1

        if "Cinema" in wb.sheetnames:
            ws = wb["Cinema"]
            rows = list(ws.iter_rows(values_only=True))
            for row in rows[1:]:
                if not any(row):
                    continue
                try:
                    nome     = _val(row[1])
                    città    = _val(row[2])
                    num_sale = int(row[3]) if row[3] else 1
                    telefono = _val(row[4])
                    indirizzo = _val(row[5])
                    lat = _f(row[6])
                    lng = _f(row[7])
                except Exception:
                    continue
                if not nome:
                    continue
                if nome in existing_nomi:
                    skipped_cinemas += 1
                    continue
                self.create_cinema(nome=nome, città=città, num_sale=num_sale,
                                   telefono=telefono, indirizzo=indirizzo, lat=lat, lng=lng)
                existing_nomi.add(nome)
                added_cinemas += 1

        return {
            "added_problems": added_problems,
            "skipped_problems": skipped_problems,
            "added_cinemas": added_cinemas,
            "skipped_cinemas": skipped_cinemas,
        }


# ── Istanza globale ───────────────────────────────────
store = ExcelStore()
