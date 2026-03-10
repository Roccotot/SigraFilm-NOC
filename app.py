from flask import Flask, render_template, request, redirect, url_for, session, flash, abort, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from storage import store

# --- CONFIGURAZIONE ---
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "devsecret-change-me")

# Inizializza file Excel e dati di default
store.seed()
print("📂 DATABASE: file Excel in cartella data/")


# --- ROUTES ---
@app.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


# --- LOGIN ---
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        u = store.get_user_by_username(username)
        if u and check_password_hash(u.password_hash, password):
            session["user_id"] = u.id
            session["role"] = u.role
            session["username"] = u.username
            flash("Login effettuato", "success")
            return redirect(url_for("dashboard"))
        flash("Credenziali non valide", "danger")
    return render_template("login.html")


# --- RESET ADMIN (emergenza) ---
@app.route("/reset-admin-password-7x9k")
def reset_admin_password():
    u = store.get_user_by_username("admin")
    if u:
        u.password_hash = generate_password_hash("admin1234")
        u.password_plain = "admin1234"
        u.role = "admin"
        store.update_user(u)
        return "Password admin resettata a 'admin1234'."
    store.create_user("admin", generate_password_hash("admin1234"), "admin1234", "admin")
    return "Utente admin ricreato con password 'admin1234'."


# --- LOGOUT ---
@app.route("/logout")
def logout():
    session.clear()
    flash("Logout effettuato", "info")
    return redirect(url_for("login"))


# --- DASHBOARD ---
@app.route("/dashboard")
def dashboard():
    if "user_id" not in session:
        return redirect(url_for("login"))

    filter_urgenza = request.args.get("filter_urgenza", "")
    filter_stato   = request.args.get("filter_stato", "")
    uid = session["user_id"]

    problems = store.get_problems_filtered(
        stato_ne="Chiuso",
        autore=session["username"] if session["role"] != "admin" else None,
        urgenza=filter_urgenza or None,
        stato_eq=filter_stato or None,
    )

    all_open = store.get_problems_filtered(
        stato_ne="Chiuso",
        autore=session["username"] if session["role"] != "admin" else None,
    )
    stats = {
        "total":    len(all_open),
        "aperto":   sum(1 for p in all_open if p.stato == "Aperto"),
        "in_corso": sum(1 for p in all_open if p.stato == "In corso"),
        "chiuso":   0,
        "critico":  sum(1 for p in all_open if p.urgenza == "Critico"),
    }

    if session["role"] == "admin":
        cinemas = store.get_all_cinemas()
    else:
        assigned_ids = store.get_cinema_ids_for_user(uid)
        cinemas = store.get_cinemas_by_ids(assigned_ids) if assigned_ids else store.get_all_cinemas()
    cinemas.sort(key=lambda c: c.nome)
    single_cinema = cinemas[0] if len(cinemas) == 1 else None

    # Contatori messaggi non letti
    reads = store.get_reads_by_user(uid)
    chat_info = {}
    for p in problems:
        comments = store.get_comments(p.id)
        total = len(comments)
        last_read = reads.get(p.id)
        if last_read is None:
            unread = total
        else:
            unread = sum(1 for c in comments if c.data_ora and c.data_ora > last_read)
        chat_info[p.id] = {"total": total, "unread": unread}

    return render_template(
        "dashboard.html",
        problems=problems,
        filter_urgenza=filter_urgenza,
        filter_stato=filter_stato,
        stats=stats,
        cinemas=cinemas,
        chat_info=chat_info,
        single_cinema=single_cinema,
    )


# --- DETTAGLIO TICKET ---
@app.route("/problems/<int:problem_id>", methods=["GET"])
def ticket_detail(problem_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    p = store.get_problem_by_id(problem_id)
    if not p:
        abort(404)
    if session["role"] != "admin" and session["username"] != p.autore:
        return "Accesso negato", 403
    comments = store.get_comments(p.id)
    store.upsert_ticket_read(session["user_id"], p.id)
    return render_template("ticket_detail.html", problem=p, comments=comments)


# --- AGGIUNGI COMMENTO ---
@app.route("/problems/<int:problem_id>/comment", methods=["POST"])
def add_comment(problem_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    p = store.get_problem_by_id(problem_id)
    if not p:
        abort(404)
    if session["role"] != "admin" and session["username"] != p.autore:
        return "Accesso negato", 403
    testo = request.form.get("testo", "").strip()
    if testo:
        store.add_comment(p.id, session["username"], session["role"], testo)
    return redirect(url_for("ticket_detail", problem_id=p.id) + "#chat-bottom")


# --- AGGIORNA TICKET (stato/urgenza) ---
@app.route("/problems/<int:problem_id>/update", methods=["POST"])
def update_ticket(problem_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    p = store.get_problem_by_id(problem_id)
    if not p:
        abort(404)
    if session["role"] != "admin" and session["username"] != p.autore:
        return "Accesso negato", 403
    nuovo_stato   = request.form.get("stato", p.stato)
    nuova_urgenza = request.form.get("urgenza", p.urgenza)
    if nuovo_stato == "Chiuso" and p.stato != "Chiuso":
        p.chiuso_da = session["username"]
        p.chiuso_il = datetime.utcnow()
    elif nuovo_stato != "Chiuso":
        p.chiuso_da = None
        p.chiuso_il = None
    p.stato   = nuovo_stato
    p.urgenza = nuova_urgenza
    store.update_problem(p)
    flash("Ticket aggiornato.", "success")
    if nuovo_stato == "Chiuso":
        return redirect(url_for("closed_tickets"))
    return redirect(url_for("ticket_detail", problem_id=p.id))


# --- ARCHIVIO TICKET CHIUSI ---
@app.route("/closed")
def closed_tickets():
    if "user_id" not in session:
        return redirect(url_for("login"))
    problems = store.get_problems_filtered(
        stato_eq="Chiuso",
        autore=session["username"] if session["role"] != "admin" else None,
    )
    return render_template("closed_tickets.html", problems=problems)


# --- AGGIUNGI PROBLEMA ---
@app.route("/problems/add", methods=["POST"])
def add_problem():
    if "user_id" not in session:
        return redirect(url_for("login"))
    cinema_nome = request.form.get("cinema", "").strip()
    sala    = request.form.get("sala", "1").strip()
    tipo    = request.form.get("tipo", "").strip()
    urgenza = request.form.get("urgenza", "Non urgente")
    stato   = request.form.get("stato", "Aperto")
    if not cinema_nome or not tipo:
        flash("Compila tutti i campi.", "danger")
        return redirect(url_for("dashboard"))
    cinema_obj = store.get_cinema_by_nome(cinema_nome)
    città = cinema_obj.città if cinema_obj else ""
    store.create_problem(cinema=cinema_nome, città=città, sala=sala,
                         tipo=tipo, urgenza=urgenza, stato=stato,
                         autore=session["username"])
    flash("Problema aggiunto con successo.", "success")
    return redirect(url_for("dashboard"))


# --- MODIFICA PROBLEMA ---
@app.route("/problems/<int:problem_id>/edit", methods=["GET", "POST"])
def edit_problem(problem_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    p = store.get_problem_by_id(problem_id)
    if not p:
        abort(404)
    if session["role"] != "admin" and session["username"] != p.autore:
        return "Accesso negato", 403
    if request.method == "POST":
        p.cinema  = request.form.get("cinema", p.cinema)
        p.tipo    = request.form.get("tipo", p.tipo)
        p.urgenza = request.form.get("urgenza", p.urgenza)
        p.stato   = request.form.get("stato", p.stato)
        store.update_problem(p)
        flash("Problema aggiornato con successo.", "success")
        return redirect(url_for("dashboard"))
    cinemas = store.get_all_cinemas()
    return render_template("edit_problem.html", problem=p, cinemas=cinemas)


# --- ARCHIVIA PROBLEMA ---
@app.route("/problems/<int:problem_id>/delete", methods=["POST"])
def delete_problem(problem_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    p = store.get_problem_by_id(problem_id)
    if not p:
        abort(404)
    if session["role"] != "admin" and session["username"] != p.autore:
        return "Accesso negato", 403
    p.stato = "Chiuso"
    store.update_problem(p)
    flash("Ticket archiviato.", "success")
    return redirect(url_for("dashboard"))


# --- ELIMINA DEFINITIVAMENTE (solo admin) ---
@app.route("/problems/<int:problem_id>/destroy", methods=["POST"])
def destroy_problem(problem_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    if session["role"] != "admin":
        return "Accesso negato", 403
    p = store.get_problem_by_id(problem_id)
    if not p:
        abort(404)
    store.delete_problem(problem_id)
    flash("Ticket eliminato definitivamente.", "success")
    return redirect(url_for("closed_tickets"))


# --- GESTIONE UTENTI ---
@app.route("/users", methods=["GET", "POST"])
def admin_users():
    if session.get("role") != "admin":
        return "Accesso negato", 403
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        role     = request.form.get("role", "user")
        telefono = request.form.get("telefono", "").strip()
        email    = request.form.get("email", "").strip()
        if not username or len(password) < 8:
            flash("Username obbligatorio e password di almeno 8 caratteri.", "danger")
            return redirect(url_for("admin_users"))
        if store.get_user_by_username(username):
            flash("Username già in uso.", "warning")
            return redirect(url_for("admin_users"))
        store.create_user(username=username, password_hash=generate_password_hash(password),
                          password_plain=password, role=role, telefono=telefono, email=email)
        flash("Utente creato con successo.", "success")
        return redirect(url_for("admin_users"))
    users_list = sorted(store.get_all_users(), key=lambda u: u.id)
    return render_template("users.html", users=users_list)


# --- DETTAGLIO UTENTE ---
@app.route("/users/<int:user_id>", methods=["GET", "POST"])
def user_detail(user_id):
    if session.get("role") != "admin":
        return "Accesso negato", 403
    u = store.get_user_by_id(user_id)
    if not u:
        abort(404)
    if request.method == "POST":
        cinema_ids = [int(x) for x in request.form.getlist("cinema_ids") if x.isdigit()]
        store.set_user_cinemas(u.id, cinema_ids)
        flash(f"Cinema assegnati a '{u.username}' aggiornati.", "success")
        return redirect(url_for("user_detail", user_id=u.id))
    all_cinemas  = store.get_all_cinemas(order_by="città_nome")
    assigned_ids = set(store.get_cinema_ids_for_user(u.id))
    return render_template("user_detail.html", u=u, all_cinemas=all_cinemas, assigned_ids=assigned_ids)


# --- RESET PASSWORD ---
@app.route("/users/<int:user_id>/reset", methods=["POST"])
def reset_password(user_id):
    if session.get("role") != "admin":
        return "Accesso negato", 403
    new_password = request.form.get("new_password", "").strip()
    if len(new_password) < 8:
        flash("La nuova password deve avere almeno 8 caratteri.", "danger")
        return redirect(url_for("admin_users"))
    u = store.get_user_by_id(user_id)
    if not u:
        abort(404)
    u.password_hash  = generate_password_hash(new_password)
    u.password_plain = new_password
    store.update_user(u)
    flash(f"Password di '{u.username}' aggiornata con successo.", "success")
    return redirect(url_for("admin_users"))


# --- ELIMINA UTENTE ---
@app.route("/users/<int:user_id>/delete", methods=["POST"])
def delete_user(user_id):
    if session.get("role") != "admin":
        return "Accesso negato", 403
    if session.get("user_id") == user_id:
        flash("Non puoi eliminare il tuo stesso utente mentre sei loggato.", "warning")
        return redirect(url_for("admin_users"))
    u = store.get_user_by_id(user_id)
    if not u:
        abort(404)
    if u.role == "admin" and store.count_admins() <= 1:
        flash("Non puoi eliminare l'unico admin rimasto.", "warning")
        return redirect(url_for("admin_users"))
    username = u.username
    store.delete_user(user_id)
    flash(f"Utente '{username}' eliminato.", "success")
    return redirect(url_for("admin_users"))


# --- GESTIONE CINEMA ---
@app.route("/admin/cinemas", methods=["GET", "POST"])
def admin_cinemas():
    if session.get("role") != "admin":
        return "Accesso negato", 403
    if request.method == "POST":
        nome     = request.form.get("nome", "").strip()
        città    = request.form.get("città", "").strip()
        telefono = request.form.get("telefono", "").strip()
        indirizzo= request.form.get("indirizzo", "").strip()
        try:
            num_sale = max(1, int(request.form.get("num_sale", "1")))
        except ValueError:
            num_sale = 1
        if nome:
            store.create_cinema(nome=nome, città=città, num_sale=num_sale,
                                telefono=telefono, indirizzo=indirizzo)
            flash(f"Cinema '{nome}' ({città}) aggiunto.", "success")
        return redirect(url_for("admin_cinemas"))

    cinemas = store.get_all_cinemas(order_by="città_nome")
    _urgency_order = {"Critico": 0, "Urgente": 1, "Non urgente": 2}
    open_problems = store.get_problems_filtered(stato_ne="Chiuso")
    open_problems.sort(key=lambda p: _urgency_order.get(p.urgenza, 9))
    tickets_map = {}
    for p in open_problems:
        key = (p.cinema or "").strip()
        if key:
            tickets_map.setdefault(key, []).append(p)
    return render_template("cinemas.html", cinemas=cinemas, tickets_map=tickets_map)


@app.route("/admin/cinemas/<int:cinema_id>/edit", methods=["GET", "POST"])
def edit_cinema(cinema_id):
    if session.get("role") != "admin":
        return "Accesso negato", 403
    c = store.get_cinema_by_id(cinema_id)
    if not c:
        abort(404)
    if request.method == "POST":
        nuovo_nome  = request.form.get("nome", "").strip()
        nuova_città = request.form.get("città", "").strip()
        try:
            num_sale = max(1, int(request.form.get("num_sale", "1")))
        except ValueError:
            num_sale = 1
        try:
            lat = float(request.form.get("lat", "").strip()) if request.form.get("lat", "").strip() else None
            lng = float(request.form.get("lng", "").strip()) if request.form.get("lng", "").strip() else None
        except ValueError:
            lat = lng = None
        if nuovo_nome:
            c.nome     = nuovo_nome
            c.città    = nuova_città
            c.num_sale = num_sale
            c.telefono = request.form.get("telefono", "").strip()
            c.indirizzo= request.form.get("indirizzo", "").strip()
            c.lat      = lat
            c.lng      = lng
            store.update_cinema(c)
            flash(f"Cinema '{nuovo_nome}' aggiornato.", "success")
        return redirect(url_for("admin_cinemas"))
    return render_template("edit_cinema.html", c=c)


@app.route("/admin/cinemas/<int:cinema_id>/delete", methods=["POST"])
def delete_cinema(cinema_id):
    if session.get("role") != "admin":
        return "Accesso negato", 403
    c = store.get_cinema_by_id(cinema_id)
    if c:
        nome = c.nome
        store.delete_cinema(cinema_id)
        flash(f"Cinema '{nome}' eliminato.", "success")
    return redirect(url_for("admin_cinemas"))


# --- EXPORT EXCEL ---
@app.route("/export/excel")
def export_excel():
    if "user_id" not in session:
        return redirect(url_for("login"))

    is_admin = session["role"] == "admin"
    username = session["username"]
    foglio   = request.args.get("foglio", "tutto")

    wb = openpyxl.Workbook()
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1F2937")
    center_align = Alignment(horizontal="center", vertical="center")

    def style_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align

    def autowidth(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    def fmt(dt):
        return dt.strftime("%d/%m/%Y %H:%M") if dt else ""

    first_sheet = True

    def new_sheet(title):
        nonlocal first_sheet
        if first_sheet:
            ws = wb.active
            ws.title = title
            first_sheet = False
        else:
            ws = wb.create_sheet(title)
        return ws

    if foglio in ("aperti", "tutto"):
        ws = new_sheet("Ticket Aperti")
        q = store.get_problems_filtered(stato_ne="Chiuso",
                                        autore=username if not is_admin else None)
        style_header(ws, ["ID", "Cinema", "Città", "Sala", "Descrizione", "Urgenza", "Stato", "Autore", "Data apertura"])
        for p in q:
            ws.append([p.id, p.cinema, p.città, p.sala, p.tipo, p.urgenza, p.stato, p.autore, fmt(p.data_ora)])
        autowidth(ws)

    if foglio in ("chiusi", "tutto"):
        ws = new_sheet("Archivio Chiusi")
        q2 = store.get_problems_filtered(stato_eq="Chiuso",
                                         autore=username if not is_admin else None)
        style_header(ws, ["ID", "Cinema", "Città", "Sala", "Descrizione", "Urgenza", "Autore", "Data apertura", "Chiuso da", "Chiuso il"])
        for p in q2:
            ws.append([p.id, p.cinema, p.città, p.sala, p.tipo, p.urgenza, p.autore, fmt(p.data_ora), p.chiuso_da or "", fmt(p.chiuso_il)])
        autowidth(ws)

    if foglio in ("cinema", "tutto") and is_admin:
        ws = new_sheet("Cinema")
        style_header(ws, ["ID", "Nome", "Città", "Sale", "Telefono", "Indirizzo", "Lat", "Lng"])
        for c in store.get_all_cinemas(order_by="città_nome"):
            ws.append([c.id, c.nome, c.città, c.num_sale, c.telefono, c.indirizzo, c.lat or "", c.lng or ""])
        autowidth(ws)

    if foglio in ("utenti", "tutto") and is_admin:
        ws = new_sheet("Utenti")
        style_header(ws, ["ID", "Username", "Ruolo", "Email", "Telefono"])
        for u in sorted(store.get_all_users(), key=lambda x: x.id):
            ws.append([u.id, u.username, u.role, u.email, u.telefono])
        autowidth(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    now = datetime.now().strftime("%Y%m%d_%H%M")
    nomi = {"aperti": "ticket_aperti", "chiusi": "archivio_chiusi",
            "cinema": "cinema", "utenti": "utenti", "tutto": "completo"}
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"sigrafilm_{nomi.get(foglio, foglio)}_{now}.xlsx",
    )


# --- IMPORT EXCEL ---
@app.route("/import/excel", methods=["GET", "POST"])
def import_excel():
    if "user_id" not in session:
        return redirect(url_for("login"))
    if session["role"] != "admin":
        return "Accesso negato", 403
    if request.method == "GET":
        return render_template("import_excel.html")

    f = request.files.get("file")
    if not f or not f.filename.endswith(".xlsx"):
        flash("Carica un file .xlsx valido.", "danger")
        return redirect(url_for("import_excel"))
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception:
        flash("File non valido o corrotto.", "danger")
        return redirect(url_for("import_excel"))

    counts = store.import_from_workbook(wb)
    parts = []
    if counts["added_problems"]:   parts.append(f"{counts['added_problems']} ticket aggiunti")
    if counts["skipped_problems"]: parts.append(f"{counts['skipped_problems']} ticket già presenti (saltati)")
    if counts["added_cinemas"]:    parts.append(f"{counts['added_cinemas']} cinema aggiunti")
    if counts["skipped_cinemas"]:  parts.append(f"{counts['skipped_cinemas']} cinema già presenti (saltati)")
    if not parts:
        flash("Nessuna nuova riga trovata — tutto già presente.", "info")
    else:
        flash(" · ".join(parts) + ".", "success")
    return redirect(url_for("import_excel"))


# --- ERRORE 500 ---
@app.errorhandler(500)
def _internal_error(e):
    flash("Errore temporaneo del server. Riprova.", "warning")
    return redirect(request.referrer or url_for("dashboard"))


# --- MAIN ---
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
